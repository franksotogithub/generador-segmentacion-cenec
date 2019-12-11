# -*-coding: utf-8-*-

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl import drawing
import win32com.client
from os import path
import string
import xlsxwriter
from conf import config
from datetime import datetime


escudo_nacional = path.join(config.PATH_IMG,'Escudo_BN.png')
logo_nacional = path.join(config.PATH_IMG,'Inei_BN.png')

FILAS_DEFAULT =34
FILA_INICIAL = 14
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

titleFill = PatternFill(start_color='D9D9D9',
                        end_color='D9D9D9',
                        fill_type='solid')

SHEEP='programacion'
SHEEP_ETIQUETAS ='etiquetas'
TIPO_FORMATO =1

def abc(inicial, final):
    abc = list(string.ascii_uppercase)
    ini = abc.index(inicial)
    fin = abc.index(final)
    lista = abc[ini:fin + 1]
    return lista


def set_border(ws):

    list_col = abc('B', 'E') + abc('K', 'O')
    for i in list_col:
        ws["{}11".format(i)].border = thin_border
        ws["{}12".format(i)].border = thin_border
    list_fil = abc('E','G')
    for i in list_fil:
        ws["{}6".format(i)].border = thin_border
        ws["{}7".format(i)].border = thin_border
        ws["{}8".format(i)].border = thin_border
    list_fil2 = abc('F', 'M')
    for i in list_fil2:
        ws["{}10".format(i)].border = thin_border


def insertImage(ws):
    img = drawing.image.Image(escudo_nacional)
    img.drawing.width = 90
    img.drawing.height = 90
    ws.add_image(img, "C1")
    img2 = drawing.image.Image(logo_nacional)
    img2.drawing.width = 90
    img2.drawing.height = 90
    ws.add_image(img2, "X1")

def cuerpo_border(ws, filas):
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 4
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 13
    ws.column_dimensions['N'].width = 13
    ws.column_dimensions['O'].width = 5
    ws.column_dimensions['P'].width = 10
    ws.column_dimensions['Q'].width = 10
    ws.column_dimensions['R'].width = 10
    ws.column_dimensions['S'].width = 11
    ws.column_dimensions['T'].width = 10
    ws.column_dimensions['U'].width = 13
    ws.column_dimensions['V'].width = 13
    ws.column_dimensions['W'].width = 10
    ws.column_dimensions['X'].width = 13
    ws.column_dimensions['Y'].width = 10
    ws.row_dimensions[12].height = 15
    ws.row_dimensions[13].height = 35

    columns = range(2,26)


    rows = range(12, filas+3)
    for col in columns:
        for row in rows:
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            ws.cell(row=row, column=col).font = Font(bold=True, size=10)

    #rows = range(12, 37)
    #for col in columns:
    #    for row in rows:
    #        ws.cell(row=row, column=col).border = thin_border
    #        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
    #        ws.cell(row=row, column=col).font = Font(bold=True, size=10)

def cuerpo_sede_border(ws, filas):
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 10
    ws.column_dimensions['M'].width = 15
    ws.column_dimensions['N'].width = 4
    ws.column_dimensions['O'].width = 10
    ws.column_dimensions['P'].width = 13
    ws.column_dimensions['Q'].width = 13
    ws.column_dimensions['R'].width = 5
    ws.column_dimensions['S'].width = 10
    ws.column_dimensions['T'].width = 10
    ws.column_dimensions['U'].width = 10
    ws.column_dimensions['V'].width = 11
    ws.column_dimensions['W'].width = 10
    ws.column_dimensions['X'].width = 13
    ws.column_dimensions['Y'].width = 13
    ws.column_dimensions['Z'].width = 10
    ws.column_dimensions['AA'].width = 13
    ws.column_dimensions['AB'].width = 10
    ws.row_dimensions[12].height = 15
    ws.row_dimensions[13].height = 35

    columns = range(2,29)


    rows = range(12, filas+3)
    for col in columns:
        for row in rows:
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
            ws.cell(row=row, column=col).font = Font(bold=True, size=10)

def colorCelda(ws):
    ws['N6'].fill = titleFill
    for row in range(6,11):
        ws['B{}'.format(row)].fill = titleFill

    for x in abc('B', 'Y'):
        ws['{}12'.format(x)].fill = titleFill
        ws['{}13'.format(x)].fill = titleFill

    for x in range(34,36):
        ws['B{}'.format(x)].fill = titleFill

def diasCuerpo(ws):
    n = 0
    lista = abc('B', 'P')
    for x in lista:
        n = n + 1
        ws["{}10".format(x)] = u'{}'.format(n)
        ws["{}10".format(x)].alignment = Alignment(horizontal="center", vertical="center")

def formatoCamposDecimal(ws):
    m = 0
    n = 0
    lista1 = abc('U', 'Y')
    lista2 = list(range(FILA_INICIAL, 36))
    for p in lista1:
        for q in lista2:
            ws["{}{}".format(p, q)].number_format = '0.00'


def llenarTotal(ws):
    for x in abc('F', 'O'):
        ws['{}13'.format(x)] = "=SUM({}14: {}100)".format(x, x)

def cabecera(data_cabecera,wb,filas):
    cod_oper=data_cabecera['COD_OPER']
    ws = wb.get_sheet_by_name(SHEEP)

    ws["D1"] = u'V CENSO NACIONAL ECONÓMICO 2019 - 2020'
    ws["D2"] = u'2019 - 2020'
    ws["D4"] = u'PROGRAMACIÓN DE RUTA DEL EMPADRONADOR'


    ws["B6"] = u'A. ORGANIZACION DE CAMPO '
    ws["B7"] = u'SEDE OPERATIVA'
    ws["E7"] = data_cabecera['CODSEDE']
    ws["F7"] = data_cabecera['SEDE_OPERATIVA']
    if TIPO_FORMATO == '90':
        ws["B8"] = u'BRIGADA'
    else:
        ws["B8"] = u'EQUIPO'

    ws["E8"] = data_cabecera['BRIGADA']

    if 'RUTA' in data_cabecera.keys() :

        if TIPO_FORMATO == 1:
            ws["B9"] = u'RUTA'
            ws["E9"] = data_cabecera['RUTA']
            ws["D4"] = u'PROGRAMACIÓN DE RUTA DEL EMPADRONADOR'
            ws["B10"] = u'EMPADRONADOR'
            ws["E10"] = data_cabecera['EMPADRONADOR']
            ws["N10"] = u'DOC.CENEC.03.11'
            ws["N6"] = u'B. NOMBRE Y APELLIDO DEL EMPADRONADOR'
        else:
            ws["B9"] = u'RUTA'
            ws["E9"] = data_cabecera['RUTA']
            ws["D4"] = u'PROGRAMACIÓN DE RUTA DEL/ DE LA REGISTRADOR/A'
            ws["B10"] = u'REGISTRADOR/A'
            ws["E10"] = data_cabecera['EMPADRONADOR']
            ws["N10"] = u'DOC.CENEC.03.23'
            ws["N6"] = u'B. NOMBRE Y APELLIDO DEL REGISTRADOR/A'


    else:
        if TIPO_FORMATO == 1:
            ws["D4"] = u'PROGRAMACIÓN DE RUTA DEL JEFE DE BRIGADA'
            ws["N10"] = u'DOC.CENEC.03.12'
            ws["N6"] = u'B. NOMBRE Y APELLIDO DEL JEFE DE BRIGADA'
        else:
            ws["D4"] = u'PROGRAMACIÓN DE RUTA DEL/DE LA JEFE/A DE EQUIPO'
            ws["N10"] = u'DOC.CENEC.03.24'
            ws["N6"] = u'B. NOMBRE Y APELLIDO DEL/DE LA JEFE/A DE EQUIPO'

    ws["N7"] = u''


    ws["N10"].alignment = Alignment(horizontal="right", vertical="bottom")
    ws["N10"].font = Font(bold=True)
    ws["B12"] = u'N° ORD'
    ws["C12"] = u'UBIGEO'
    ws["D12"] = u'DEPARTAMENTO'
    ws["E12"] = u'PROVINCIA'
    ws["F12"] = u'DISTRITO'
    ws["G12"] = u'COD CCPP'
    ws["H12"] = u'CENTRO POBLADO'
    ws["I12"] = u'UBICACIÓN CENSAL'   #MERGE
    ws["K12"] = u'EST.'
    ws["L12"] = u'N° PERIODO'
    ws["M12"] = u'PERIODO TRABAJO'    #MERGE
    ws["O12"] = u'DÍAS DE OPERACIÓN DE CAMPO' #MERGE
    ws["U12"] = u'ASOCIACIÓN DE FONDOS' #MERGE

    ws["I13"] = u'ZONA'
    ws["J13"] = u'MANZANA'
    ws["M13"] = u'FECHA INI'
    ws["N13"] = u'FECHA FIN'
    ws["O13"] = u'VIAJE'
    ws["P13"] = u'DIAS TRABAJO'
    ws["Q13"] = u'GABINETE'
    ws["R13"] = u'DESCANSO'
    ws["S13"] = u'DÍAS OPERATIVOS'
    ws["T13"] = u'TOT.DÍAS'
    ws["U13"] = u'TOTAL MOV.LOCAL'
    ws["V13"] = u'TOTAL MOV.ESPECIAL'
    ws["W13"] = u'PASAJES'
    ws["X13"] = u'TOTAL VIÁTICOS'
    ws["Y13"] = u'TOTAL GENERAL s/.'

    ws["B{}".format(filas)] = u'TOTAL DEL PERIODO'
    ws["B{}".format(filas+1)] = u'TOTAL DE LA RUTA'
    ws["B{}".format(filas+2)] = u'OBSERVACIONES'



    celdasCabecera = [
        'D1:W1','D2:W2','D4:W4',
        'B12:B13',
        'C12:C13',
        'D12:D13',
        'E12:E13',
        'F12:F13',
        'G12:G13',
        'H12:H13',
        'I12:J12',
        'K12:K13',
        'L12:L13',
        'M12:N12',
        'O12:T12',
        'U12:Y12',
        'N10:Y10',
        'N6:Y6',
        'N7:Y7',
        'B6:F6',
        'B7:D7',
        'B8:D8',
        'B9:D9',
        'B10:D10',
        'E8:F8',
        'E9:F9',
        'E10:F10',
        'B{row}:J{row}'.format(row=filas),
        'B{row}:J{row}'.format(row=filas + 1),
        'B{row}:J{row}'.format(row=filas + 2),
        'K{row}:Y{row}'.format(row=filas + 2),

    ]

    for cells in celdasCabecera:
        ws.merge_cells(cells)

    list = abc('O', 'Y')



    for x in list:
        ws['{column}{row}'.format(column=x ,row=filas)] = "=SUM({column}{ini}: {column}{fin})".format(column = x,ini = FILA_INICIAL ,fin = filas-1)
        ws['{column}{row}'.format(column=x,row=filas+1)] = "=SUM({column}{ini}: {column}{fin})".format(column = x,ini = FILA_INICIAL ,fin = filas-1)



    cells_cabecera=[]

    for row in range(12,14):
        for column in abc('B', 'Y'):
            cell= '{}{}'.format(column,row)
            ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[cell].font = Font(bold=True, size=8)
            ws[cell].border = thin_border

    for row in range(6,11):
        ws['B{}'.format(row)].alignment = Alignment( vertical="center", wrap_text=True)
        ws['B{}'.format(row)].font = Font(bold=True, size=12)
        for column in abc('B', 'F'):
            ws['{}{}'.format(column,row)].border = thin_border


    for row in range(6, 8):
        for column in abc('N', 'Y'):
            ws['{}{}'.format(column, row)].border = thin_border


    list_cell = ['D1','D2','D4','B6','N6','B34','B35','B36']
    for cell in list_cell:
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell not in ['B6','N6','B34','B35','B36']:
            ws[cell].font = Font(bold=True, size=16)
        else:
            ws[cell].font = Font(bold=True, size=12)


def cabecera_sede(data_cabecera,wb,filas):

    ws = wb.get_sheet_by_name(SHEEP)

    ws["D1"] = u'V CENSO NACIONAL ECONÓMICO 2019 - 2020'
    ws["D2"] = u'2019 - 2020'
    ws["D4"] = u'PROGRAMACIÓN DE RUTAS DE EMPADRONADOR POR SEDE'

    #ws["D4"] = u'PROGRAMACIÓN DE RUTAS DE EMPADRONADOR POR SEDE'

    ws["B6"] = u'A. ORGANIZACION DE CAMPO '
    ws["B7"] = u'SEDE OPERATIVA'
    ws["D7"] = data_cabecera['CODSEDE']
    ws["E7"] = data_cabecera['SEDE_OPERATIVA']

    ws["N6"] = u'B. NOMBRE Y APELLIDO'
    ws["N7"] = u''


    ws["N10"].alignment = Alignment(horizontal="right", vertical="bottom")
    ws["N10"].font = Font(bold=True)
    ws["B12"] = u'N° ORD'
    ws["C12"] = u'BRIGADA'
    ws["D12"] = u'RUTA'
    ws["E12"] = u'EMP'
    ws["F12"] = u'UBIGEO'
    ws["G12"] = u'DEPARTAMENTO'
    ws["H12"] = u'PROVINCIA'
    ws["I12"] = u'DISTRITO'
    ws["J12"] = u'COD CCPP'
    ws["K12"] = u'CENTRO POBLADO'
    ws["L12"] = u'UBICACIÓN CENSAL'   #MERGE
    ws["N12"] = u'EST.'
    ws["O12"] = u'N° PERIODO'
    ws["P12"] = u'PERIODO TRABAJO'    #MERGE
    ws["R12"] = u'DÍAS DE OPERACIÓN DE CAMPO' #MERGE
    ws["X12"] = u'ASOCIACIÓN DE FONDOS' #MERGE

    ws["L13"] = u'ZONA'
    ws["M13"] = u'MANZANA'
    ws["P13"] = u'FECHA INI'
    ws["Q13"] = u'FECHA FIN'
    ws["R13"] = u'VIAJE'
    ws["S13"] = u'DIAS TRABAJO'
    ws["T13"] = u'GABINETE'
    ws["U13"] = u'DESCANSO'
    ws["V13"] = u'DÍAS OPERATIVOS'
    ws["W13"] = u'TOT.DÍAS'
    ws["X13"] = u'TOTAL MOV.LOCAL'
    ws["Y13"] = u'TOTAL MOV.ESPECIAL'
    ws["Z13"] = u'PASAJES'
    ws["AA13"] = u'TOTAL VIÁTICOS'
    ws["AB13"] = u'TOTAL GENERAL s/.'

    ws["B{}".format(filas)] = u'TOTAL DEL PERIODO'
    ws["B{}".format(filas+1)] = u'TOTAL DE LA RUTA'
    ws["B{}".format(filas+2)] = u'OBSERVACIONES'

    celdasCabecera = [
        'D1:W1','D2:W2','D4:W4',
        'B12:B13',
        'C12:C13',
        'D12:D13',
        'E12:E13',
        'F12:F13',
        'G12:G13',
        'H12:H13',
        'I12:I13',
        'J12:J13',
        'K12:K13',
        'L12:M12',
        'N12:N13',
        'O12:O13',
        'P12:Q12',
        'R12:W12',
        'X12:AB12',
        'N10:AB10',
        'N6:AB6',
        'N7:AB7',
        'B6:H6',
        'B7:C7',
        'E7:H7',
        'B{row}:M{row}'.format(row=filas),
        'B{row}:M{row}'.format(row=filas+1),
        'B{row}:M{row}'.format(row=filas+2),
        'N{row}:AB{row}'.format(row=filas+2),
    ]

    for cells in celdasCabecera:
        ws.merge_cells(cells)

    list = abc('R', 'Z')
    list.extend(['AA', 'AB'])


    for x in list:
        ws['{column}{row}'.format(column=x ,row=filas)] = "=SUM({column}{ini}: {column}{fin})".format(column = x,ini = FILA_INICIAL ,fin = filas-1)
        ws['{column}{row}'.format(column=x,row=filas+1)] = "=SUM({column}{ini}: {column}{fin})".format(column = x,ini = FILA_INICIAL ,fin = filas-1)

    cells_cabecera=[]

    for row in range(12,FILA_INICIAL):
        list = abc('B', 'Z')
        list.extend(['AA', 'AB'])
        for column in list:
            cell= '{}{}'.format(column,row)
            ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws[cell].font = Font(bold=True, size=8)
            ws[cell].border = thin_border

    for row in range(6,8):
        ws['B{}'.format(row)].alignment = Alignment( vertical="center", wrap_text=True)
        ws['B{}'.format(row)].font = Font(bold=True, size=12)
        for column in abc('B', 'H'):
            ws['{}{}'.format(column,row)].border = thin_border


    for row in range(6, 8):
        list = abc('N', 'Z')
        list.extend(['AA', 'AB'])
        for column in list:
            ws['{}{}'.format(column, row)].border = thin_border


    list_cell = ['D1', 'D2', 'D4', 'B6', 'N6']
    for cell in list_cell:
        ws[cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if cell not in ['B6', 'N6']:
            ws[cell].font = Font(bold=True, size=16)
        else:
            ws[cell].font = Font(bold=True, size=12)


def crear_excel(cabecera, output,alto):
    wb = xlsxwriter.Workbook(path.join(config.PATH_PROGRAMACIONES,output))
    ws = wb.add_worksheet(SHEEP)
    ws.print_area(1, 2, 36, 26)
    ws.fit_to_pages(1, 2)
    wb.close()


def crear_excel_etiquetas( output):
    wb = xlsxwriter.Workbook(output)
    ws = wb.add_worksheet(SHEEP_ETIQUETAS)
    ws.print_area(1, 2, 36, 26)
    ws.fit_to_pages(1, 1)
    wb.close()

def excel2PDF(output,sheep=SHEEP,orientacion=1):
    o = win32com.client.Dispatch("Excel.Application")
    o.Visible = 0
    o.DisplayAlerts = 0
    name,formato=output.split('.')
    path_pdf = '{}{}'.format(name,'.pdf')
    wb = o.Workbooks.Open(output)
    ws=wb.Sheets(sheep)
    ws.PageSetup.Orientation=orientacion
    ws.PageSetup.LeftMargin = 1
    ws.PageSetup.RightMargin = 1
    ws.ExportAsFixedFormat(0, path_pdf)
    wb.Close()

def llenar_excel(info, output, alto):
    data_cabecera = info[0]
    wb = load_workbook(output)

    filas = FILAS_DEFAULT
    if(FILA_INICIAL+alto > FILAS_DEFAULT):
        filas = FILA_INICIAL+alto

    cabecera(data_cabecera, wb,filas)
    ws = wb.get_sheet_by_name(SHEEP)
    data_cuerpo = info[1]

    for row in range(0,alto):
        i=row + FILA_INICIAL
        x=data_cuerpo[row]
        ws["B{}".format(i)]=u'{}'.format(row+1)
        ws["C{}".format(i)]=u'{}'.format(x['UBIGEO'])
        ws["D{}".format(i)]=u'{}'.format(x['DEPARTAMENTO'])
        ws["E{}".format(i)]=u'{}'.format(x['PROVINCIA'])
        ws["F{}".format(i)]=u'{}'.format(x['DISTRITO'])
        ws["G{}".format(i)]=u'{}'.format(x['CODCCPP'])
        ws["H{}".format(i)]=u'{}'.format(x['NOMCCPP'])
        ws["I{}".format(i)]=u'{}'.format(x['ZONA'])
        ws["J{}".format(i)]=u'{}'.format(x['MANZANAS'])
        ws["K{}".format(i)]=u'{}'.format(x['TOTAL_EST'])
        ws["L{}".format(i)]=u'{}'.format(x['PERIODO'])
        ws["M{}".format(i)]=u'{}'.format((datetime.strptime(x['FECHA_INI'][0:10], "%Y-%m-%d")).strftime("%d/%m/%Y"))
        ws["N{}".format(i)]=u'{}'.format((datetime.strptime(x['FECHA_FIN'][0:10], "%Y-%m-%d")).strftime("%d/%m/%Y"))
        ws["O{}".format(i)]=int(x['DIAS_VIAJE'])
        ws["P{}".format(i)]=int(x['DIAS_TRABAJO'])
        ws["Q{}".format(i)]=int(x['GABINETE'])
        ws["R{}".format(i)]=int(x['DIAS_DESCANSO'])
        ws["S{}".format(i)]=int(x['DIAS_OPERATIVOS'])
        ws["T{}".format(i)]=int(x['TOTAL_DIAS'])
        ws["U{}".format(i)]=float(x['TOTAL_MOV_LOCAL'])
        ws["V{}".format(i)]=float(x['TOTAL_MOV_ESPECIAL'])
        ws["W{}".format(i)]=float(x['PASAJES'])
        ws["X{}".format(i)]=float(x['TOTAL_VIATICOS'])
        ws["Y{}".format(i)]=float(x['TOTAL_GENERAL'])

    insertImage(ws)
    cuerpo_border(ws, filas)
    formatoCamposDecimal(ws)


    #### text de color gris
    ws['N6'].fill = titleFill

    for row in range(6,8):
        ws['B{}'.format(row)].fill = titleFill

    list3 =abc('A', 'Y')

    for x in list3:
        ws['{}12'.format(x)].fill = titleFill
        ws['{}13'.format(x)].fill = titleFill

    for x in range(filas,filas+3):
        ws['B{}'.format(x)].fill = titleFill


    wb.save(output)
    wb.close()

def llenar_excel_sede(info, output, alto):
    data_cabecera = info[0]
    wb = load_workbook(output)

    filas = FILAS_DEFAULT
    if(FILA_INICIAL+alto > FILAS_DEFAULT):
        filas = FILA_INICIAL+alto

    cabecera_sede(data_cabecera, wb ,filas)
    ws = wb.get_sheet_by_name(SHEEP)
    data_cuerpo = info[1]




    for row in range(0,alto):
        i=row + FILA_INICIAL
        x=data_cuerpo[row]

        ws["B{}".format(i)]=u'{}'.format(row+1)

        ws["C{}".format(i)] = str(x['BRIGADA']).zfill(3)

        ws["D{}".format(i)] = str(x['RUTA']).zfill(4)

        emp = str(x['RUTA']).zfill(4)
        if int(x['ADICIONAL'])>0:
            emp='A'+str(x['RUTA']).zfill(3)

        ws["E{}".format(i)] = u'{}'.format(emp)
        ws["F{}".format(i)]=u'{}'.format(x['UBIGEO'])
        ws["G{}".format(i)]=u'{}'.format(x['DEPARTAMENTO'])
        ws["H{}".format(i)]=u'{}'.format(x['PROVINCIA'])
        ws["I{}".format(i)]=u'{}'.format(x['DISTRITO'])
        ws["J{}".format(i)]=u'{}'.format(x['CODCCPP'])
        ws["K{}".format(i)]=u'{}'.format(x['NOMCCPP'])
        ws["L{}".format(i)]=u'{}'.format(x['ZONA'])
        ws["M{}".format(i)]=u'{}'.format(x['MANZANAS'])
        ws["N{}".format(i)]=u'{}'.format(x['TOTAL_EST'])
        ws["O{}".format(i)]=u'{}'.format(x['PERIODO'])
        ws["P{}".format(i)]=u'{}'.format((datetime.strptime(x['FECHA_INI'][0:10], "%Y-%m-%d")).strftime("%d/%m/%Y"))
        ws["Q{}".format(i)]=u'{}'.format((datetime.strptime(x['FECHA_FIN'][0:10], "%Y-%m-%d")).strftime("%d/%m/%Y"))
        ws["R{}".format(i)]=int(x['DIAS_VIAJE'])
        ws["S{}".format(i)]=int(x['DIAS_TRABAJO'])
        ws["T{}".format(i)]=int(x['GABINETE'])
        ws["U{}".format(i)]=int(x['DIAS_DESCANSO'])
        ws["V{}".format(i)]=int(x['DIAS_OPERATIVOS'])
        ws["W{}".format(i)]=int(x['TOTAL_DIAS'])
        ws["X{}".format(i)]=float(x['TOTAL_MOV_LOCAL'])
        ws["Y{}".format(i)]=float(x['TOTAL_MOV_ESPECIAL'])
        ws["Z{}".format(i)]=float(x['PASAJES'])
        ws["AA{}".format(i)]=float(x['TOTAL_VIATICOS'])
        ws["AB{}".format(i)]=float(x['TOTAL_GENERAL'])

    insertImage(ws)

    cuerpo_sede_border(ws, filas)

    #####formato decimal
    lista1 = abc('X', 'Z')
    lista1.extend(['AA','AB'])
    lista2 = list(range(FILA_INICIAL, filas+3))
    for p in lista1:
        for q in lista2:
            ws["{}{}".format(p, q)].number_format = '0.00'




    #### text de color gris

    ws['N6'].fill = titleFill
    for row in range(6,8):
        ws['B{}'.format(row)].fill = titleFill

    list3 =abc('B', 'Z')
    list3.extend(['AA', 'AB'])
    for x in list3:
        ws['{}12'.format(x)].fill = titleFill
        ws['{}13'.format(x)].fill = titleFill

    for x in range(filas,filas+3):
        ws['B{}'.format(x)].fill = titleFill


    wb.save(output)
    wb.close()


def llenar_excel_etiqueta(c, ws,x):

    i=x

    cod_oper=c['COD_OPER']



    ws["B{}".format(i)] = u'Sede'

    ws["C{}".format(i)] = u'{}'.format(c['SEDE_OPERATIVA'])

    if TIPO_FORMATO == 1:
        ws["D{}".format(i)] = u'Brigada'
    else:
        ws["D{}".format(i)] = u'Equipo'


    ws["E{}".format(i)] = u'{}'.format(c['BRIGADA'])
    i = i+1
    if 'RUTA' in c.keys():
        ws["B{}".format(i)] = u'Ruta'
        if TIPO_FORMATO ==1:

            ws["D{}".format(i)] = u'Empadronador'

        else:
            ws["D{}".format(i)] = u'Registrador'

        ws["C{}".format(i)] = u'{}'.format(c['RUTA'])
        ws["E{}".format(i)] = u'{}'.format(c['EMPADRONADOR'])
        i = i + 1

    ws["B{}".format(i)] =u'Periodo'
    ws["D{}".format(i)] = u'{}'.format(c['PERIODO'])
    i = i + 1
    ws["B{}".format(i)].font =Font(name='3 of 9 Barcode' ,size=36)
    ws["B{}".format(i)].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if 'RUTA' in c.keys():
        ws["B{}".format(i)] = '*{cod_oper}{periodo}{codsede}{brigada}{ruta}{emp}*'.format(cod_oper=c['COD_OPER'],
                                                                               periodo = c['PERIODO'],
                                                                               codsede=c['CODSEDE'],
                                                                               brigada=c['BRIGADA'], ruta=c['BRIGADA'],
                                                                               emp=c['EMPADRONADOR'])

    else:
        ws["B{}".format(i)] = '*{cod_oper}{periodo}{codsede}{brigada}*'.format(cod_oper=c['COD_OPER'],
                                                                    periodo=c['PERIODO'],
                                                                    codsede=c['CODSEDE'],
                                                                    brigada=c['BRIGADA'])

    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    merge_celdas = [
        'B{i}:C{i}'.format(i=i-1),
        'D{i}:E{i}'.format(i=i-1),
        'B{i}:E{i}'.format(i=i),
    ]

    for cells in merge_celdas:
        ws.merge_cells(cells)


    lista1 = abc('B', 'E')
    lista2 = list(range(x, i+1))
    for p in lista1:
        for q in lista2:
            ws["{}{}".format(p, q)].border = thin_border
            if p in ['B','D'] and q<i:
                ws["{}{}".format(p, q)].fill = titleFill
                ws["{}{}".format(p, q)].font = Font(bold=True)
            elif p in ['B','D'] and (q==i):
                ws["{}{}".format(p, q)].alignment = Alignment(horizontal="center", vertical="center")



    i = i + 2

    return i

def crear_programacion_rutas(info,output,tipo_formato=1):
    TIPO_FORMATO = tipo_formato
    cabecera=info[0]
    cuerpo = info[1]
    alto = len(cuerpo)
    crear_excel(cabecera,output,alto)
    llenar_excel(info,output,alto)
    excel2PDF(output,sheep=SHEEP,orientacion=2)

def crear_programacion_brigadas(info,output,tipo_formato=1):
    TIPO_FORMATO = tipo_formato
    cabecera = info[0]
    cuerpo = info[1]
    alto = len(cuerpo)
    crear_excel(cabecera, output, alto)
    llenar_excel(info, output, alto)
    excel2PDF(output,sheep=SHEEP,orientacion=2)


def crear_etiquetas(data,output,tipo_formato=1):
    TIPO_FORMATO = tipo_formato
    crear_excel_etiquetas(output)

    j=2
    wb = load_workbook(output)
    ws = wb.get_sheet_by_name(SHEEP_ETIQUETAS)
    for e in data:
        j=llenar_excel_etiqueta(e, ws,j)
    wb.save(output)
    wb.close()
    excel2PDF(output,sheep=SHEEP_ETIQUETAS)
    #for ruta in rutas ()




def crear_programacion_sedes(info,output):
    cabecera = info[0]
    cuerpo = info[1]
    alto = len(cuerpo)
    crear_excel(cabecera, output, alto)
    llenar_excel_sede(info, output, alto)
    excel2PDF(output,sheep=SHEEP,orientacion=2)